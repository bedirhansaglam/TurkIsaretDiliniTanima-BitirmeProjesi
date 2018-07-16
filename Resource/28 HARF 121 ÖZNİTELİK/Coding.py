# -*- coding: utf-8 -*-
from PyQt4 import QtGui
from PyQt4 import QtCore
import numpy as np
from tasarim import Ui_Dialog
import confusion
import x_train
import x_test
import y_test_table
import y_train_table
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.metrics import accuracy_score
from sklearn.metrics import confusion_matrix
from DataBase import connect
from DataBase import selectTable
from sklearn.neural_network import MLPClassifier
from sklearn.externals import joblib
from sklearn.tree import DecisionTreeClassifier
from sklearn.naive_bayes import GaussianNB
from openpyxl import Workbook

    
class MainWindow(QtGui.QMainWindow, Ui_Dialog):

    def __init__(self):
        QtGui.QMainWindow.__init__(self)
        self.setupUi(self)
        self.statusBar().showMessage(unicode("Hazir\n")) 
        self.btn_dosyaYukle3.clicked.connect(self.VerileriYukle)
        self.hs_test_value.valueChanged.connect(self.hs_valueChanged)
        self.cb_classificier.currentIndexChanged.connect(self.cb_classificier_changed)
        self.btn_split.clicked.connect(self.train_and_test_click)
        self.liste=[]
        self.cnx=connect("root","","127.0.0.1","yirmisekizharf") #Veritabanı baglanti ayarlari
        self.secim=0
        self.modelName='Random Forest'
        self.confusionMatrix=confusion.Ui_Dialog() #confusionMatrix Örneği alınıyor
        self.confusionWidget = QtGui.QWidget() 
        self.confusionMatrix.setupUi(self.confusionWidget) # veriler widgeta set ediliyor.
        
        #Eğitim verilerini göstermek için form örneği oluşturuluyor.
        self.formXtrain=x_train.Ui_Dialog()
        self.formXtrainWidget=QtGui.QWidget()
        self.formXtrain.setupUi(self.formXtrainWidget)
        #Eğitim verilerinin etiketleri gösterilmek için form örneği oluşturuluyor
        self.formYtrain=y_train_table.Ui_Dialog()
        self.formYtrainWidget=QtGui.QWidget()
        self.formYtrain.setupUi(self.formYtrainWidget)
        
        #Test verilerini göstermek için form örneği oluşturuluyor
        self.formXtest=x_test.Ui_Dialog()
        self.formXtestWidget=QtGui.QWidget()
        self.formXtest.setupUi(self.formXtestWidget)
        
        #Test Verilerinin etiketleri gösterilmek için form örneği oluşturuluyor
        self.formYtest=y_test_table.Ui_Dialog()
        self.formYtestWidget=QtGui.QWidget()
        self.formYtest.setupUi(self.formYtestWidget)
        #Butonlar aktif ediliyor
        self.btn_confusion.clicked.connect(self.confusion_show)
        self.pb_x_train.clicked.connect(self.x_train_show)
        self.pb_x_test.clicked.connect(self.x_test_show)
        self.pb_y_test.clicked.connect(self.y_test_show)
        self.pb_y_train.clicked.connect(self.y_train_show)
    
    def confusion_show(self):
        self.confusionWidget.show()
    def x_train_show(self):
        self.formXtrainWidget.show()
    def x_test_show(self):
        self.formXtestWidget.show()
    def y_test_show(self):
        self.formYtestWidget.show()
    def y_train_show(self):
        self.formYtrainWidget.show()
    
    def VerileriYukle(self): #veritabanına bağlanıp tablodan veri çekme işlemleri
        cursor=self.cnx.cursor(buffered=True)
        self.alldata=[]
        select=selectTable(cursor)
        
        for a in select:
            self.alldata.append(a)
        self.alldata=np.array(self.alldata)
        self.verileri_isle(self.alldata,self.table_all_data)
        veri_sayisi=self.alldata.shape[0]
        oznitelik_sayisi=self.alldata.shape[1]
        self.statusBar().showMessage(unicode("Veriler Yüklendi...Toplam "+str(veri_sayisi+1)+" veri ve "+str(oznitelik_sayisi)+" öznitelik bulunuyor.",'utf-8'))
 
    def verileri_yaz(self,veri,isim):
        harfler=['A','B','C',unicode('Ç','utf-8'),'D','E','F','G','H','I',unicode('İ','utf-8'),'J','K','L','M','N','O',unicode('Ö','utf-8'),'P','R','S',unicode('Ş','utf-8'),'T','U',unicode('Ü','utf-8'),'V','Y','Z']
        excel=Workbook()
        excelws=excel['Sheet']

        for i,harf in enumerate(harfler):
            excelws.cell(row=0,column=i+1).value=harf
            excelws.cell(row=i+1,column=0).value=harf
            
        for rowNumber,row in enumerate(veri): #Veri satır satır okunuyor
            for columnNumber in range(0,len(veri[0])): # Her bir satırdaki veri sütun sütun okunuyor
                excelws.cell(row=rowNumber+1, column=columnNumber+1).value = str(row[columnNumber]) #tablonun rowNumber'ıncı satır columnNumber'ıncı sütununa veri ekleniyor.
        excel.save(u"./"+isim+".xlsx")
        
    def verileri_dok(self,X,tablo):
        num_rows=len(X) #gönderilen veriden satır sayısı alınıyor
        tablo.clear()    # tablo temizleniyor
        tablo.setColumnCount(len(X[0])) # gönderilen veriden sütun sayısı alınıyor ve tabloya set ediliyor.
        tablo.setRowCount(num_rows) #alınan satır sayısı tabloya set ediliyor.
        for rowNumber,row in enumerate(X): #Veri satır satır okunuyor
            for columnNumber in range(0,len(X[0])): # Her bir satırdaki veri sütun sütun okunuyor
                tablo.setItem(rowNumber,columnNumber,QtGui.QTableWidgetItem(str(row[columnNumber]))) #tablonun rowNumber'ıncı satır columnNumber'ıncı sütununa veri ekleniyor.
    
    def test_verilerini_dok(self,X,tablo):
        num_rows=len(X) #gönderilen veriden satır sayısı alınıyor
        tablo.clear()    # tablo temizleniyor
        tablo.setColumnCount(1) # gönderilen veriden sütun sayısı alınıyor ve tabloya set ediliyor.
        tablo.setRowCount(num_rows) #alınan satır sayısı tabloya set ediliyor.
        for rowNumber,row in enumerate(X): #Veri satır satır okunuyor
            tablo.setItem(rowNumber,0,QtGui.QTableWidgetItem(str(row))) #tablonun rowNumber'ıncı satır columnNumber'ıncı sütununa veri ekleniyor
                
    def verileri_isle(self,X,tablo):
        islenmisveri=[]   
        for rowNumber,row in enumerate(X): #Veritabanından çekilen veriler işlenmek üzere satır satır okunuyor
            temp=[] #geçiçi bir liste oluşturuluyor
            listx=[] #1.el için kemiklerin x koordinatlarını tutabilmek için liste tanımlanıyor.
            listy=[] #1.el için kemiklerin y koordinatlarını tutabilmek için liste tanımlanıyor.
            listz=[] #1.el için kemiklerin z koordinatlarını tutabilmek için liste tanımlanıyor.
            for i in range(0,3): #0-1-2 indisler 1.el için palm_position değerlerini tutmakta
                for j in range(3,63): #3-62 arasındaki indisler 1.el için m_e , p_e,i_e,d_e değerlerini tutmakta
                    if(i==0 and j%3==0): # mod 3 alarak kemiklerin x,y,z koordinatlarını elde ediyoruz.
                        a=row[i]-row[j] #(pp_x)-(m_e_x) ,(pp_x)-(p_e_x) , (pp_x)-(i_e_x) , (pp_x)-(d_e_x) işlemleri sırasıyla yapılır
                        listx.append(a) # daha sonra bu değerler listede tutulur.
                    elif (i==1 and j%3==1):
                        a=row[i]-row[j] #(pp_y)-(m_e_y) ,(pp_y)-(p_e_y) , (pp_y)-(i_e_y) , (pp_y)-(d_e_y) işlemleri sırasıyla yapılır
                        listy.append(a) # daha sonra bu değerler listede tutulur.
                    elif (i==2 and j%3==2):
                        a=row[i]-row[j] #(pp_z)-(m_e_z) ,(pp_z)-(p_e_z) , (pp_z)-(i_e_z) , (pp_z)-(d_e_z) işlemleri sırasıyla yapılır
                        listz.append(a) # daha sonra bu değerler listede tutulur.
                    
            for i in range(0,20):   #listelerde tutulan x,y,z değerleri doğru yerlerine yerleştirilmek için farklı 
                temp.append(listx[i])#bir dizide doğru sıra ile toplanırlar
                temp.append(listy[i])
                temp.append(listz[i])
                
            #yukarıda 1.el için yapılan işlemler , 2. el için yapılır.
            listx=[]
            listy=[]
            listz=[]
            for i in range(63,66):
                for j in range(66,126):
                    if(i==63 and j%3==0):
                        a=row[i]-row[j]
                        tablo.setItem(rowNumber,j,QtGui.QTableWidgetItem(str(a)))
                        listx.append(a)
                    elif (i==64 and j%3==1):
                        a=row[i]-row[j]
                        tablo.setItem(rowNumber,j,QtGui.QTableWidgetItem(str(a)))
                        listy.append(a)
                    elif (i==65 and j%3==2):
                        a=row[i]-row[j]
                        tablo.setItem(rowNumber,j,QtGui.QTableWidgetItem(str(a)))
                        listz.append(a)
                        
            for i in range(0,20):
                temp.append(listx[i])
                temp.append(listy[i])
                temp.append(listz[i])
            
            temp.append(row[126])  #Yapılan işlemlerin sonunda işlenmiş bir satır veriye en son etiketi eklenir ve bu veri
            islenmisveri.append(temp) #işlenmiş olarak yeni listede yerini alır.
        islenmisveri=np.array(islenmisveri) # işlenmiş veriler dizi ye çevrilir
        self.verileri_dok(islenmisveri,tablo) # diziye çevrilen veri kullanıcıya gösterilmek üzere tabloya gönderilir.
        self.alldata=islenmisveri #işlenen veriyi sınıflandırma işlemlerinde kullanılması için global bir değişkene aktarılır.
    

    def cb_classificier_changed(self,value): # sınıflandırıcı türünü seçmek için kullanılan combobox 'ın changeIndex Metodu.
        if (value==0):
            self.secim=0
            self.modelName='RandomForestModel'
        elif(value==1):
            self.secim=1
            self.modelName='NeuralNetworkModel'
        elif(value==2):
            self.secim=2
            self.modelName='GauisanNavieBayesModel'
        elif(value==3):
            self.secim=3
            self.modelName='SupportVectorMachineModel'
        elif(value==4):
            self.secim=4
            self.modelName='DecisionTreeModel'
            
    def hs_valueChanged(self): # Eğitim ve test verisini ayırırken test verisinin yüzdesini belirlemek amacıyla kullanılan horizontal slider valueChanged olayı
        val="% "+str(self.hs_test_value.value()) #Value nun görsel olarak kullanıcılara aktarılması
        self.lbl_hs_value.setText(val)
        
    def train_and_test_click(self):
        X=self.alldata[:,:120]
        y=self.alldata[:,120]
        value=float(self.hs_test_value.value())
        value=value/100
        X_train,X_test,y_train,y_test=train_test_split(X,y,test_size=value)
        self.classification(X_train,y_train,X_test,y_test)
        
    def classification(self,X_train,y_train,X_test,y_test):
        if(self.secim==0): #randomForest
            clf=RandomForestClassifier(max_depth=None,random_state=42)
        elif(self.secim==1): #neuralNetwork
            clf=MLPClassifier(activation='tanh', solver='adam', alpha=1e-5,hidden_layer_sizes=(5,), random_state=1)     
        elif(self.secim==2):#decisionTree
            clf=DecisionTreeClassifier()
        elif(self.secim==3):#supportVectorMachine
            clf=SVC()
        elif(self.secim==4):#gauisanNavieBayes
            clf=GaussianNB()
        
        clf.fit(X_train,y_train) #sınıflandırıcı veriler ile eğitiliyor
        mname='./models/'+str(self.modelName)+'.pkl' #eğitilen modele isim veriliyor
        joblib.dump(clf,mname ) #eğitilen model kayıt ediliyor
        result=clf.predict(X_test) # Model üzerinde test verileri test ediliyor
        self.verileri_dok(confusion_matrix(result,y_test,labels=[1,2,3,4,5,6,7,8,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29]),self.confusionMatrix.table_confusion_matrix) #doğru sonuçlar ile elde edilen sonuçlar karşılaştırılıyor ve karmaşıklık matrixi oluşturuluyor.
        self.confusionMatrix.table_confusion_matrix.setHorizontalHeaderLabels(['A','B','C',unicode('Ç','utf-8'),'D','E','F','G','H','I',unicode('İ','utf-8'),'J','K','L','M','N','O',unicode('Ö','utf-8'),'P','R','S',unicode('Ş','utf-8'),'T','U',unicode('Ü','utf-8'),'V','Y','Z'])
        self.confusionMatrix.table_confusion_matrix.setVerticalHeaderLabels(['A','B','C',unicode('Ç','utf-8'),'D','E','F','G','H','I',unicode('İ','utf-8'),'J','K','L','M','N','O',unicode('Ö','utf-8'),'P','R','S',unicode('Ş','utf-8'),'T','U',unicode('Ü','utf-8'),'V','Y','Z'])
        lbl_text=("Basari orani: %"+ str ((round(accuracy_score(y_test,result),2))*100)) # doğru sonuçlar ile elde edilen sonuçlar karşılaştırılıyor ve başarı oranı ölçülüyor. Kullanıcıya gösteriliyor.
        self.lbl_basari.setText(lbl_text) #kullanıcıya gösterilmek için label set ediliyor.
        
        self.verileri_dok(X_train,self.formXtrain.table_x_train) #Eğitim Verileri gösterilmek için set ediliyor
        self.verileri_dok(X_test,self.formXtest.table_x_test) # Test Verileri gösterilmek için set ediliyor
        self.test_verilerini_dok(y_train,self.formYtrain.table_y_train) #Eğitim verilerinin etikleri gösteriliyor
        self.test_verilerini_dok(y_test,self.formYtest.table_y_test) #Test verileri etiketleri gösteriliyor
        
        self.verileri_yaz(confusion_matrix(result,y_test,labels=[1,2,3,4,5,6,7,8,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29]),self.modelName+" Confusion Matrix")
        
        


          
